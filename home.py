import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from sklearn.cluster import KMeans
from sklearn.cluster import AgglomerativeClustering

def show():
    st.title("Home Page")
    st.write("Upload an Excel file to optimize your data.")

    uploaded_file = st.file_uploader("Choose an Excel file", type=["xlsx", "xls"])
    if uploaded_file is not None:
        try:
            # Read from defined table name using openpyxl
            wb = load_workbook(uploaded_file, data_only=True)

# # Entry Bulan April
#             with st.expander("Entry Bulan April", expanded=False):
#                 ws = wb.worksheets[0]  # Use the first worksheet
#                 start_row = 3
#                 end_row = ws.max_row
#                 min_col = 1  # Column A
#                 max_col = 31  # Column AE
#                 data = ws.iter_rows(min_row=start_row, max_row=end_row, min_col=min_col, max_col=max_col, values_only=True)
#                 data = list(data)
#                 if len(data) > 1:
#                     # Exclude rows where the first column is 'JUMLAH'
#                     filtered_data = [row for row in data[1:] if row[0] != 'JUMLAH']
#                     df = pd.DataFrame(filtered_data, columns=data[0])
#                 else:
#                     df = pd.DataFrame()
#                 st.dataframe(df, use_container_width=True, hide_index=True)


#             with st.expander("Entry Detail Stunting", expanded=False):
#                 ws = wb.worksheets[1]
#                 start_row = 3
#                 end_row = ws.max_row
#                 min_col = 1
#                 max_col = 34
#                 data = ws.iter_rows(min_row=start_row, max_row=end_row, min_col=min_col, max_col=max_col, values_only=True)
#                 data = list(data)
#                 if len(data) > 1:
#                     filtered_data = [row for row in data[1:] if row[0] != 'JUMLAH']
#                     df = pd.DataFrame(filtered_data, columns=data[0])
#                 else:
#                     df = pd.DataFrame()
#                 st.dataframe(df, use_container_width=True, hide_index=True)

#             with st.expander("Entry By Kelurahan", expanded=False):
#                 ws = wb.worksheets[2]
#                 start_row = 3
#                 end_row = ws.max_row
#                 min_col = 1
#                 max_col = 21
#                 data = ws.iter_rows(min_row=start_row, max_row=end_row, min_col=min_col, max_col=max_col, values_only=True)
#                 data = list(data)
#                 if len(data) > 1:
#                     filtered_data = [row for row in data[1:] if row[0] != 'JUMLAH']
#                     df = pd.DataFrame(filtered_data, columns=data[0])
#                 else:
#                     df = pd.DataFrame()
#                 st.dataframe(df, use_container_width=True, hide_index=True)

#             with st.expander("Ranking By Kelurahan", expanded=False):
#                 ws = wb.worksheets[3]
#                 start_row = 3
#                 end_row = ws.max_row
#                 min_col = 1
#                 max_col = 3
#                 data = ws.iter_rows(min_row=start_row, max_row=end_row, min_col=min_col, max_col=max_col, values_only=True)
#                 data = list(data)
#                 if len(data) > 1:
#                     filtered_data = [row for row in data[1:] if row[0] != 'JUMLAH']
#                     df = pd.DataFrame(filtered_data, columns=data[0])
#                 else:
#                     df = pd.DataFrame()
#                 st.dataframe(df, use_container_width=True, hide_index=True)

#             with st.expander("Entry By Kecamatan", expanded=False):
#                 ws = wb.worksheets[4]
#                 start_row = 3
#                 end_row = ws.max_row
#                 min_col = 1
#                 max_col = 20
#                 data = ws.iter_rows(min_row=start_row, max_row=end_row, min_col=min_col, max_col=max_col, values_only=True)
#                 data = list(data)
#                 if len(data) > 1:
#                     filtered_data = [row for row in data[1:] if row[0] != 'JUMLAH']
#                     df = pd.DataFrame(filtered_data, columns=data[0])
#                 else:
#                     df = pd.DataFrame()
#                 st.dataframe(df, use_container_width=True, hide_index=True)

#             with st.expander("Ranking By Kecamatan", expanded=False):
#                 ws = wb.worksheets[5]
#                 start_row = 3
#                 end_row = ws.max_row
#                 min_col = 1
#                 max_col = 3
#                 data = ws.iter_rows(min_row=start_row, max_row=end_row, min_col=min_col, max_col=max_col, values_only=True)
#                 data = list(data)
#                 if len(data) > 1:
#                     filtered_data = [row for row in data[1:] if row[0] != 'JUMLAH']
#                     df = pd.DataFrame(filtered_data, columns=data[0])
#                 else:
#                     df = pd.DataFrame()
#                 st.dataframe(df, use_container_width=True, hide_index=True)

            st.success("##### Chart Rank by Kelurahan")
            ws = wb.worksheets[4]
            start_row = 3
            end_row = ws.max_row
            min_col = 1  # No column
            max_col = 3  # Total data column
            data = ws.iter_rows(min_row=start_row, max_row=end_row, min_col=min_col, max_col=max_col, values_only=True)
            data = list(data)
            if len(data) > 1:
                # Stop processing when a row with column 1 value 'JUMLAH' is found
                filtered_data = []
                for row in data[1:]:
                    if row[0] == 'JUMLAH':
                        break
                    filtered_data.append(row)
                if filtered_data:
                    df_chart = pd.DataFrame(filtered_data, columns=data[0])
                    name_col = df_chart.columns[1]
                    total_col = df_chart.columns[2]
                    st.bar_chart(df_chart.set_index(name_col)[[total_col]])

                    # --- KMeans Clustering Example ---
                    st.info("KMeans Clustering on Total Data By Kelurahan")
                    X = df_chart[[total_col]].astype(float)
                    n_clusters = st.slider("Number of clusters for KMeans", 1, 100, 3)
                    kmeans = KMeans(n_clusters=n_clusters, random_state=42)
                    kmeans.fit(X)
                    df_chart['Cluster'] = kmeans.labels_
                    col1, col2 = st.columns([1, 1], gap="large")
                    with col1:
                        st.dataframe(df_chart[[name_col, total_col, 'Cluster']], use_container_width=True)
                    with col2:
                        st.dataframe(pd.DataFrame(kmeans.cluster_centers_, columns=[total_col]), use_container_width=True)

                    # --- Agglomerative Clustering Example ---
                    st.info("Agglomerative Clustering on Total Data By Kelurahan")
                    n_clusters_agg = st.slider("Number of clusters for Agglomerative", 2, 6, 3, key="agglo")
                    agglo = AgglomerativeClustering(n_clusters=n_clusters_agg)
                    agglo_labels = agglo.fit_predict(X)
                    df_chart['AggloCluster'] = agglo_labels
                    col3, col4 = st.columns([1, 1], gap="large")
                    with col3:
                        st.dataframe(df_chart[[name_col, total_col, 'AggloCluster']], use_container_width=True)
                    with col4:
                        st.write("Agglomerative does not provide cluster centers.")
                else:
                    st.warning("Not enough data to display chart.")
            else:
                st.warning("Not enough data to display chart.")

            # st.info("##### Pie Chart Rank by Kecamatan")
            # ws = wb.worksheets[5]
            # start_row = 3
            # end_row = ws.max_row
            # min_col = 1  # No column
            # max_col = 3  # Total data column
            # data = ws.iter_rows(min_row=start_row, max_row=end_row, min_col=min_col, max_col=max_col, values_only=True)
            # data = list(data)
            # if len(data) > 1:
            #     # Stop processing when a row with column 1 value 'JUMLAH' is found
            #     filtered_data = []
            #     for row in data[1:]:
            #         if row[0] == 'JUMLAH':
            #             break
            #         filtered_data.append(row)
            #     if filtered_data:
            #         df_chart = pd.DataFrame(filtered_data, columns=data[0])
            #         name_col = df_chart.columns[1]
            #         total_col = df_chart.columns[2]
            #         st.plotly_chart(
            #             {
            #                 "data": [{
            #                     "type": "pie",
            #                     "labels": df_chart[name_col],
            #                     "values": df_chart[total_col],
            #                     "hole": 0.3
            #                 }]
            #             },
            #             use_container_width=True
            #         )
            #     else:
            #         st.warning("Not enough data to display chart.")
            # else:
            #     st.warning("Not enough data to display chart.")

        except Exception as e:
            st.error(f"Error reading the Excel file: {e}")